#!/usr/bin/env python
'''Multiplication table maker
Author: Carissa Jones
'''

import argparse
import openpyxl
import sys

if len(sys.argv)!=2:
    print('Usage: python multiplicationTable.py <int>')
    sys.exit(-1)
number = None
try:
    number = abs(int(sys.argv[1]))
except ValueError as e:
    print('First argument should be an integer')
    sys.exit(-1)
wb = openpyxl.Workbook()
sheet = wb.active
boldFont = openpyxl.styles.Font(bold=True)
for rowNum in range(1, number+2):
    for colNum in range(1, number+2):
		if rowNum==1 and colNum==1:
			sheet.cell(row=rowNum, column=colNum).value=''
		elif rowNum==1:
			sheet.cell(row=rowNum, column=colNum).value = colNum-1
			sheet.cell(row=rowNum, column=colNum).font = boldFont
		elif colNum==1:
			sheet.cell(row=rowNum, column=colNum).value=rowNum-1
			sheet.cell(row=rowNum, column=colNum).font = boldFont
		else:
			sheet.cell(row=rowNum, column=colNum).value = (rowNum-1)*(colNum-1)
wb.save('table_{0}.xlsx'.format(number))
sys.exit(0)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('N', help='size of multiplication table')

    args = parser.parse_args()

    create_multiplication_table(args.N)

if __name__=='__main__':
    main()